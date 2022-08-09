from openpyxl import *
from global_instance import *
import os

class MyExcel():
    def __init__(self, *args, **kwargs):
        pass

    def excel_create(session, location, filename):
        wb = Workbook()
        wb.save(location + filename + '.xlsx')
        wb = load_workbook(location + filename + '.xlsx')
        xl_loc_dict[session] = location + filename + '.xlsx'
        xl_wb_dict[session] = wb
        ws = wb.active

    def excel_open(session, location, filename):
        wb = load_workbook(location + filename + '.xlsx')
        xl_loc_dict[session] = location + filename + '.xlsx'
        xl_wb_dict[session] = wb
        ws = wb.active

    def excel_delete_file(location, filename):
        os.remove(location + filename + '.xlsx')

    def excel_create_worksheet(session, sheetname, position, value):
        wb = xl_wb_dict[session]
        if(position == 'default'):
            wb.create_sheet(sheetname)
        elif(position == 'positional'):
            wb.create_sheet(sheetname, int(value))
        wb.save(xl_loc_dict[session])
    
    def excel_delete_worksheet(session, sheetname):
        wb = xl_wb_dict[session]
        ws = wb.get_sheet_by_name(sheetname)
        wb.remove_sheet(ws)
        wb.save(xl_loc_dict[session])
    
    def excel_save_workbook(session, location, filename):
        wb = xl_wb_dict[session]
        wb.save(location + filename + '.xlsx')
        xl_loc_dict[session] = location + filename + '.xlsx'
    
    def excel_set_value(session, sheetname, cell, value):
        wb = xl_wb_dict[session]
        ws = wb.get_sheet_by_name(sheetname)
        ws[variable_dict[cell]] = variable_dict[value]
        print(cell, value, variable_dict[value], variable_dict[cell])
        wb.save(xl_loc_dict[session])
    
    def excel_get_value(session, sheetname, cell, variable):
        wb = xl_wb_dict[session]
        ws = wb.get_sheet_by_name(sheetname)
        c = ws[cell]
        variable_dict[variable] = c.value
    
    def excel_set_formula(session, sheetname, cell, formula):
        wb = xl_wb_dict[session]
        ws = wb.get_sheet_by_name(sheetname)
        ws[cell] = formula
    
    def excel_merge_unmerge(session, sheetname, merge_unmerge, from_cell, to_cell ):
        wb = xl_wb_dict[session]
        ws = wb.get_sheet_by_name(sheetname)
        if(merge_unmerge == 'merge'):
            ws.merge_cells(from_cell +':'+ to_cell)
        elif(merge_unmerge == 'unmerge'):
            ws.unmerge_cells(from_cell +':'+ to_cell)
        wb.save(xl_loc_dict[session])

    def excel_copy_paste(copy_session, copy_sheetname, copy_start, copy_end, paste_session, paste_sheetname, paste_start, paste_end ):
        wb_copy = xl_wb_dict[copy_session]
        ws_copy = wb_copy.get_sheet_by_name(copy_sheetname)
        wb_paste = xl_wb_dict[paste_session]
        ws_paste = wb_paste.get_sheet_by_name(paste_sheetname)
        mr = ws_copy.max_row
        mc = ws_copy.max_column
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
                c = ws_copy.cell(row = i, column = j)
                ws_paste.cell(row = i, column = j).value = c.value
        wb_copy.save(xl_loc_dict[copy_session])
        wb_paste.save(xl_loc_dict[paste_session])